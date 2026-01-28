import base64
import json
import re
from io import BytesIO

import xlsxwriter
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.humanize.templatetags.humanize import intcomma
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import OuterRef, Subquery, Sum, Count, F, Q
from django.db.models.functions import Coalesce
from django.http import Http404, JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.utils import timezone
from django.utils.text import slugify
from django.utils.translation import gettext_lazy as _
from django.views.generic import ListView, DetailView, TemplateView
from django_filters.views import FilterView
from openpyxl import load_workbook

from apps.main.models import Bazaar, Area, Section
from apps.stall.filters import StallFilter
from apps.stall.forms import StallCashForm
from apps.stall.models import StallStatus, Stall
from smartbozor.qrcode import render_qr_png_file


class StallBazaarChoiceView(LoginRequiredMixin, TemplateView):
    template_name = 'main/bazaar-choice.j2'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["route"] = "stall:list"
        context["extra"] = {
            bid: _("{0} ta rasta").format(n) for bid, n in Stall.objects.annotate(
                bazaar_id=F("section__area__bazaar_id")
            ).values(
                "bazaar_id",
            ).annotate(
                n=Count("id")
            ).values_list("bazaar_id", "n")
        }

        return context


class StallListView(LoginRequiredMixin, PermissionRequiredMixin, FilterView, ListView):
    model = Stall
    paginate_by = 50
    ordering = '-id'
    template_name = 'stall/list.j2'
    permission_required = "stall.view_stall"
    filterset_class = StallFilter

    def get(self, request, *args, **kwargs):
        self.set_bazaar(request, *args, **kwargs)
        return super().get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("export", '0') == '1':
            return self.export(context)

        return super().render_to_response(context, **response_kwargs)

    def post(self, request, *args, **kwargs):
        self.set_bazaar(request, *args, **kwargs)

        if not self.bazaar.is_allow_cash or not request.user.has_perm("stall.add_stallstatus"):
            return JsonResponse({
                "success": False,
                "message": "Permission denied",
            })

        data = StallCashForm(data=self.request.POST)
        if not data.is_valid():
            return JsonResponse({
                "success": False,
                "error": data.errors,
            })

        try:
            with transaction.atomic():
                today = timezone.localtime().date()
                stall = Stall.objects.select_for_update().select_related("section").filter(
                    section__area__bazaar_id=self.bazaar.id
                ).get(pk=data.cleaned_data["stall"].id)

                if not self.bazaar.is_working_day:
                    raise Exception(_('Bugun bozor ish kuni emas'))

                ss, created = StallStatus.objects.get_or_create(
                    stall=stall,
                    date=today,
                    defaults={
                        'is_occupied': True,
                        'is_paid': True,
                        'payment_method': Bazaar.PAYMENT_METHOD_CASH,
                        'payment_progress': 0,
                        'price': stall.price,
                        'occupied_at': timezone.now(),
                        'paid_at': timezone.now(),
                    }
                )

                if not created:
                    if ss.is_paid:
                        raise Exception(_("Rasta uchun allqachon to'lov qabul qilingan"))

                    if ss.payment_progress > 0:
                        raise Exception(_("{} orqali to'lov qabul qilinmoqda ...").format(ss.payment_progress_title))

                    if not ss.is_occupied:
                        ss.is_occupied = True
                        ss.occupied_at = timezone.now()

                    ss.is_paid = True
                    ss.payment_method = Bazaar.PAYMENT_METHOD_CASH
                    ss.price = stall.price
                    ss.paid_at = timezone.now()
                    ss.save()

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e),
            })

        return JsonResponse({
            "success": True,
            "message": _("Rasta uchun naqd to'lov muvaffaqiyatli qabul qilindi")
        })

    def patch(self, request, *args, **kwargs):
        if not request.user.has_perm("stall.view_stallstatus"):
            raise PermissionDenied

        self.set_bazaar(request, *args, **kwargs)

        try:
            stall_id = int(request.GET.get("id"))

            if not Stall.objects.filter(
                    section__area__bazaar_id=self.bazaar.id,
                    id=stall_id
            ).exists():
                raise Http404
        except ValueError:
            return JsonResponse({
                "success": False,
                "error": "Invalid stall id",
            })

        data = []

        for row in StallStatus.objects.filter(stall_id=stall_id, is_paid=True).order_by("-date", "-paid_at").all():
            data.append([
                f"{timezone.localtime(row.paid_at):%d.%m.%Y %H:%M}",
                Bazaar.PAYMENT_METHOD_DICT.get(row.payment_method, "-"),
                intcomma(row.price).replace(",", " ") + " " + _("so'm")
            ])

        return JsonResponse({
            "title": _("Rasta uchun to'lovlar"),
            "headers": [
                _("Sana"),
                _("To'lov turi"),
                _("Summa")
            ],
            "data": data
        })

    def put(self, request, *args, **kwargs):
        try:
            stall_id = int(request.body)
        except ValueError:
            return JsonResponse({
                "success": False,
                "error": "Invalid body",
            })

        self.set_bazaar(request, *args, **kwargs)

        if not request.user.has_perm("stall.change_stallstatus"):
            return JsonResponse({
                "success": False,
                "message": "Permission denied",
            })

        if not self.bazaar.is_working_day:
            return JsonResponse({
                "success": False,
                "message": _('Bugun bozor ish kuni emas')
            })

        is_occupied = self.toggle_occupied(self.bazaar, stall_id)

        return JsonResponse({
            "success": True,
            "is_occupied": is_occupied,
            "title": _("band") if is_occupied else _("band emas")
        })

    def export(self, context):
        self.set_bazaar(self.request, **self.kwargs)

        xls = BytesIO()

        workbook = xlsxwriter.Workbook(xls, {'in_memory': True})
        worksheet = workbook.add_worksheet(name="Rastalar")

        header_format = workbook.add_format({
            'bg_color': '#263D54',
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_color': 'white',
            'border': 1
        })

        uzs_format = workbook.add_format({
            'num_format': '#,##0 "so\'m"',
            'align': 'right'
        })

        text_center_format = workbook.add_format({
            "align": "center",
        })

        worksheet.write(0, 0, str(_("Rasta raqami")), header_format)
        worksheet.write(0, 1, str(_("Narxi")), header_format)

        worksheet.set_column(0, 0, width=12)
        worksheet.set_column(1, 1, width=12)

        for row, stall in enumerate(Stall.objects.filter(section__area__bazaar_id=self.bazaar.id).order_by('id').all(), start=1):
            worksheet.write(row, 0, stall.number, text_center_format)
            worksheet.write(row, 1, stall.price, uzs_format)

        workbook.close()

        xls.seek(0)

        name = f"{self.bazaar}-rastalar"
        response = HttpResponse(xls, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'inline; filename="{}-{:%Y-%m-%d-%H-%M}.xlsx"'.format(
            slugify(name),
            timezone.localtime().now(),
        )

        return response

    def set_bazaar(self, request, *args, **kwargs):
        try:
            self.bazaar = Bazaar.objects.filter(
                id__in=request.user.allowed_bazaar.values_list('id', flat=True)
            ).get(pk=kwargs['pk'])  # type: Bazaar
        except Bazaar.DoesNotExist:
            raise Http404

    @classmethod
    def toggle_occupied(cls, bazaar, stall_id):
        with transaction.atomic():
            stall = Stall.objects.select_for_update().select_related("section").filter(
                section__area__bazaar_id=bazaar.id
            ).get(pk=stall_id)

            ss, created = StallStatus.objects.get_or_create(
                stall=stall,
                date=timezone.localtime().date(),
                defaults={
                    'is_occupied': True,
                    'is_paid': False,
                    'payment_method': 0,
                    'payment_progress': 0,
                    'price': stall.price,
                    'occupied_at': timezone.now(),
                    'paid_at': None,
                }
            )

            if not created and not ss.is_paid:
                ss.is_occupied = not ss.is_occupied
                ss.occupied_at = timezone.now() if ss.is_occupied else None
                ss.save()

        return ss.is_occupied

    def get_queryset(self, add_annotate=True):
        today = timezone.now().date()

        def sub(field_name):
            return StallStatus.objects.filter(
                stall_id=OuterRef('pk'),
                date=today
            ).values(field_name)[:1]

        qs = super().get_queryset().filter(
            section__area__bazaar_id=self.bazaar.id
        )

        if add_annotate:
            qs = qs.annotate(
                is_occupied_today=Coalesce(Subquery(sub("is_occupied")), False),
                is_paid_today=Coalesce(Subquery(sub("is_paid")), False),
                payment_method_today=Coalesce(Subquery(sub("payment_method")), 0),
                payment_progress_today=Coalesce(Subquery(sub("payment_progress")), 0),
            )

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["bazaar"] = self.bazaar
        context["PAGE_TITLE"] = str(self.bazaar) + " » " + _("Rastalar ro'yxati")

        if self.request.user.has_perm("stall.view_stallstatus"):
            context["total"] = {
                "count": self.get_queryset(False).count(),
                "paid": self.get_queryset(True).order_by().filter(
                    is_paid_today=True
                ).aggregate(
                    count=Count("id"),
                    total=Coalesce(Sum("price"), 0),
                ),
                "occupied_count": self.get_queryset(True).order_by().filter(
                    is_occupied_today=True
                ).count(),
                "amount": self.get_queryset(True).order_by().filter(
                    is_paid_today=True
                ).values("payment_method_today").annotate(
                    total=Coalesce(Sum("price"), 0),
                ).values("payment_method_today", "total")
            }

        return context


class StallQrCode(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Stall
    permission_required = 'stall.view_stall'

    def get_queryset(self):
        return super().get_queryset().filter(
            section__area__bazaar_id__in=self.request.user.allowed_bazaar.values_list('id', flat=True)
        )

    def render_to_response(self, context, **response_kwargs):
        obj = self.object  # type: Stall

        return render_qr_png_file(obj.qr_image_file)


class StallImportView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    TITLE = _("Rastalarni import qilish")

    model = Bazaar
    permission_required = 'stall.change_stall'
    template_name = 'stall/import.j2'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        step = 1
        try:
            step = int(request.POST.get('step', 1))
        except:
            pass

        if step == 1:
            try:
                file = request.FILES.get('file')
                workbook = load_workbook(file)
                sheet = workbook.active

                stall_wrong, duplicates, stall_skip, stall_insert, stall_update, stall_delete, import_list, can_save = self.check(
                    sheet.iter_rows(values_only=True), skip_first_line=True
                )
            except Exception as e:
                messages.error(request, str(e))
                return redirect("stall:import", self.object.id)
        elif step == 2:
            stall_list = []
            try:
                stall_list = json.loads(base64.b64decode(request.POST.get('data', "")))
            except:
                step = 1
                pass

            if step == 2:
                with transaction.atomic():
                    area = Area.objects.filter(bazaar_id=self.object.id).order_by('id').first()
                    section = Section.objects.filter(area_id=area.id).order_by('id').first()

                    stall_wrong, duplicates, stall_skip, stall_insert, stall_update, stall_delete, import_list, can_save = self.check(
                        stall_list, in_transaction=True
                    )
                    if can_save:
                        if stall_insert:
                            insert_objs = [Stall(
                                section_id=section.id,
                                number=n,
                                price=p
                            ) for n, p in stall_insert]
                            Stall.objects.bulk_create(insert_objs, batch_size=10)

                        if stall_update:
                            for stall, p in stall_update:
                                stall.price = p
                                stall.save()

                        if stall_delete:
                            Stall.objects.filter(
                                id__in=[row.id for row in stall_delete]
                            ).delete()

                        messages.success(request, _("Muvaffaqiyatli import qilindi."))
                        return redirect("stall:list", self.object.id)
                    else:
                        # Agar saqlab bo'lmasa, userga ko'rsatamiz
                        step = 1

        context = self.get_context_data(object=self.object)

        context["step"] = step
        context["stall_wrong"] = stall_wrong
        context["duplicates"] = duplicates
        context["stall_insert"] = stall_insert
        context["stall_update"] = stall_update
        context["stall_delete"] = stall_delete
        context["stall_skip"] = stall_skip
        context["import_data"] = base64.b64encode(json.dumps(import_list).encode('utf-8')).decode('utf-8')
        context["can_save"] = can_save

        return self.render_to_response(context)

    def check(self, stall_list, *, in_transaction=False, skip_first_line=False):
        stall_wrong, duplicates, stall_skip, stall_insert, stall_update, stall_delete = [], [], [], [], [], []
        stall_processed, import_list = set(), []

        qs = Stall.objects.filter(
            section__area__bazaar_id=self.object.id
        )
        if in_transaction:
            qs = qs.select_for_update()

        stall_by_number = {row.number: row for row in qs.all()}

        for n, row in enumerate(stall_list):
            if skip_first_line and n == 0:  # Skip header
                continue

            if len(row) < 2:
                raise Exception(_("Faylda kamida 2 ta ustun bo'lishi lozim"))

            number, price = map(lambda s: s.strip(), map(str, row[:2]))
            import_list.append((number, price))

            if not re.match(Stall.NUMBER_PATTERN, number) or not re.match("^[0-9]+$", price):
                stall_wrong.append((number, price))
                continue

            price = int(price)

            if number in stall_processed:
                duplicates.append((number, price))
                continue

            stall_processed.add(number)

            if number not in stall_by_number:
                # Agar DB da mavjud bo'lmasa
                stall_insert.append((number, price))
            else:
                stall = stall_by_number[number]
                del stall_by_number[number]

                if stall.price == price:
                    stall_skip.append(stall)
                else:
                    stall_update.append((stall, price))

        stall_delete.extend(stall_by_number.values())
        can_save = not stall_wrong and not duplicates and (stall_insert or stall_update or stall_delete)

        return stall_wrong, duplicates, stall_skip, stall_insert, stall_update, stall_delete, import_list, can_save
