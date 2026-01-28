import base64
import json
import re
import time
from collections import defaultdict
from io import BytesIO

import xlsxwriter
from dateutil.relativedelta import relativedelta
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Sum, F, Count
from django.db.models.functions import Coalesce, TruncMonth
from django.http import JsonResponse, Http404, HttpResponse
from django.shortcuts import redirect
from django.utils import timezone
from django.utils.text import slugify
from django.utils.translation import gettext_lazy as _
from django.views.generic import ListView, DetailView, TemplateView
from django_filters.views import FilterView
from django.utils.translation import gettext_lazy as _

from humanize import intcomma
from openpyxl import load_workbook

from apps.main.models import Bazaar, Section, Area
from apps.shop.filters import ShopFilter
from apps.shop.forms import ShopCashForm
from apps.shop.models import Shop, ShopPayment, ShopStatus
from smartbozor.helpers import uz_month
from smartbozor.qrcode import render_qr_png_file


class ShopBazaarChoiceView(LoginRequiredMixin, TemplateView):
    template_name = 'main/bazaar-choice.j2'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["route"] = "shop:list"
        context["extra"] = {
            bid: _("{0} ta magazin").format(n) for bid, n in Shop.objects.annotate(
                bazaar_id=F("section__area__bazaar_id")
            ).values(
                "bazaar_id",
            ).annotate(
                n=Count("id")
            ).values_list("bazaar_id", "n")
        }

        return context


class ShopListView(LoginRequiredMixin, PermissionRequiredMixin, FilterView, ListView):
    model = Shop
    paginate_by = 50
    ordering = '-id'
    template_name = 'shop/list.j2'
    permission_required = "shop.view_shop"
    filterset_class = ShopFilter

    def get(self, request, *args, **kwargs):
        self.set_bazaar(request, *args, **kwargs)

        return super().get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("export", '0') == '1':
            return self.export(context)

        return super().render_to_response(context, **response_kwargs)

    def post(self, request, *args, **kwargs):
        self.set_bazaar(request, *args, **kwargs)

        if not self.bazaar.is_allow_cash or not request.user.has_perm("shop.add_shoppayment"):
            return JsonResponse({
                "success": False,
                "error": "Permission denied",
            })

        data = ShopCashForm(data=self.request.POST)
        if not data.is_valid():
            return JsonResponse({
                "success": False,
                "error": data.errors
            })

        try:
            with transaction.atomic():
                shop = Shop.objects.select_for_update().select_related('section').filter(
                    section__area__bazaar_id__in=self.request.user.allowed_bazaar.values_list('id', flat=True)
                ).get(pk=data.cleaned_data["shop"].id)

                ShopPayment.objects.create(
                    shop=shop,
                    date=timezone.localtime().date(),
                    payment_method=Bazaar.PAYMENT_METHOD_CASH,
                    amount=data.cleaned_data["amount"],
                    paid_at=timezone.now(),
                )
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            })

        return JsonResponse({
            "success": True,
            "message": _("Magazin hisobiga naqd muvaffaqiyatli qo'shildi")
        })

    def patch(self, request, *args, **kwargs):
        if not request.user.has_perm("shop.view_shoppayment"):
            raise PermissionDenied

        self.set_bazaar(request, *args, **kwargs)

        try:
            shop_id = int(request.GET.get("id"))

            if not Shop.objects.filter(
                    section__area__bazaar_id=self.bazaar.id,
                    id=shop_id
            ).exists():
                raise Http404
        except ValueError:
            return JsonResponse({
                "success": False,
                "error": "Invalid shop id",
            })

        can_edit = request.user.has_perm("shop.change_shoppayment")

        data, editable = [], []
        today = timezone.localtime().date()
        for row in ShopPayment.objects.filter(
                shop_id=shop_id
        ).exclude(
            paid_at__isnull=True
        ).order_by("-date", "-paid_at").all():
            paid_at = timezone.localtime(row.paid_at)
            data.append([
                f"{paid_at:%d.%m.%Y %H:%M}",
                Bazaar.PAYMENT_METHOD_DICT.get(row.payment_method, "-"),
                intcomma(row.amount).replace(",", " ") + " " + _("so'm"),
            ])

            if can_edit:
                editable.append([
                    row.id,
                    row.amount,
                    row.payment_method == Bazaar.PAYMENT_METHOD_CASH and paid_at.year == today.year and paid_at.month == today.month
                ])

        return JsonResponse({
            "title": _("Magazin uchun to'lovlar"),
            "headers": [
                _("Sana"),
                _("To'lov turi"),
                _("Summa")
            ],
            "editable": editable,
            "data": data
        })

    def put(self, request, *args, **kwargs):
        if not request.user.has_perm("shop.change_shoppayment"):
            raise PermissionDenied

        self.set_bazaar(request, *args, **kwargs)

        try:
            data = json.loads(request.body)
            payment_id, amount = int(data.get("id", 0)), int(data.get("amount", 0))
            if payment_id <= 0:
                raise Exception("Invalid payment id")

            if amount <= 0:
                raise Exception("Invalid amount")

            today = timezone.localtime().date()
            with transaction.atomic():
                payment = ShopPayment.objects.select_for_update().filter(
                    shop__section__area__bazaar_id=self.bazaar.id
                ).exclude(
                    paid_at__isnull=True
                ).get(id=payment_id)

                if payment.payment_method != Bazaar.PAYMENT_METHOD_CASH:
                    raise Exception("Invalid payment method")

                if payment.paid_at.year != today.year or payment.paid_at.month != today.month:
                    raise Exception("Invalid payment date")

                payment.amount = amount
                payment.save()
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e),
            })

        return JsonResponse({
            "success": True,
        })

    def export(self, context):
        self.set_bazaar(self.request, **self.kwargs)

        xls = BytesIO()

        workbook = xlsxwriter.Workbook(xls, {'in_memory': True})
        worksheet = workbook.add_worksheet(name="Magazinlar")

        header_format = workbook.add_format({
            'bg_color': '#263D54',
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'font_color': 'white',
            'border': 1
        })

        uzs_format = workbook.add_format({
            'num_format': '#,##0 "so\'m"',
            'align': 'right'
        })

        number_format = workbook.add_format({
            'num_format': '#,##0',
            'align': 'right'
        })

        text_center_format = workbook.add_format({
            "align": "center",
        })

        worksheet.write(0, 0, str(_("Magazin raqami")), header_format)
        worksheet.write(0, 1, str(_("Tadbirkor")), header_format)
        worksheet.write(0, 2, str(_("Ijara narxi")), header_format)

        header_data, payment_data = self.export_payment()
        for idx, title in enumerate(header_data):
            worksheet.merge_range(0, 3 + 2 * idx, 0, 4 + 2 * idx, str(title), header_format)

        worksheet.set_column(0, 0, width=12)
        worksheet.set_column(1, 1, width=30)
        worksheet.set_column(2, 2, width=12)
        worksheet.set_column(3, 3 + 2 * len(header_data), width=12)

        for row, shop in enumerate(Shop.objects.filter(section__area__bazaar_id=self.bazaar.id).order_by('id').all(),
                                   start=1):
            worksheet.write(row, 0, shop.number, text_center_format)
            worksheet.write(row, 1, shop.owner)
            worksheet.write(row, 2, shop.rent_price, uzs_format)

            for idx, (cash, click) in enumerate(payment_data[shop.id]):
                worksheet.write(row, 3 + 2 * idx, cash, number_format)
                worksheet.write(row, 4 + 2 * idx, click, number_format)

        workbook.close()

        xls.seek(0)

        name = f"{self.bazaar}-magazinlar"
        response = HttpResponse(xls, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'inline; filename="{}-{:%Y-%m-%d-%H-%M}.xlsx"'.format(
            slugify(name),
            timezone.localtime().now(),
        )

        return response

    @classmethod
    def export_payment(cls):
        end = timezone.localtime().date().replace(day=1)
        start = end - relativedelta(months=4)
        start_ = start

        qs = ShopPayment.objects.filter(
            date__gte=start,
        ).exclude(
            paid_at__isnull=True
        ).annotate(
            month=TruncMonth('date')
        ).values("shop_id", "month", "payment_method").annotate(
            total_amount=Coalesce(Sum("amount"), 0)
        ).values_list("shop_id", "month", "payment_method", "total_amount")

        header_data = []
        while start <= end:
            header_data.append(
                f"{uz_month(start)} {start:%Y}\n" + _("Naqd | Click")
            )
            start += relativedelta(months=1)

        payment_data = defaultdict(lambda: [
            [0, 0],
            [0, 0],
            [0, 0],
            [0, 0],
            [0, 0]
        ])

        def months_between(d1, d2):
            return (d2.year - d1.year) * 12 + (d2.month - d1.month)

        for shop_id, month, pm, total in qs.all():
            idx = months_between(start_, month)
            payment_data[shop_id][idx][0 if pm == Bazaar.PAYMENT_METHOD_CASH else 1] = total

        return header_data, payment_data

    def set_bazaar(self, request, *args, **kwargs):
        try:
            self.bazaar = Bazaar.objects.filter(
                id__in=request.user.allowed_bazaar.values_list('id', flat=True)
            ).get(pk=kwargs['pk'])  # type: Bazaar
        except Bazaar.DoesNotExist:
            raise Http404

    def get_queryset(self, add_annotate=True):
        return super().get_queryset().filter(
            section__area__bazaar_id=self.bazaar.id
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["bazaar"] = self.bazaar

        shops_id = [row.id for row in context["object_list"]]

        context["total_rent_by_shop"] = {row["shop_id"]: row["amount"] for row in ShopStatus.objects.filter(
            shop_id__in=shops_id
        ).filter(is_occupied=True).values("shop_id").annotate(
            amount=Coalesce(Sum("rent_price"), 0)
        ).values("shop_id", "amount")}

        context["total_payment_by_shop"] = {row["shop_id"]: row["total_amount"] for row in ShopPayment.objects.filter(
            shop_id__in=shops_id
        ).exclude(
            paid_at__isnull=True
        ).values("shop_id").annotate(
            total_amount=Coalesce(Sum("amount"), 0)
        ).values("shop_id", "total_amount")}

        start = timezone.localtime().today().date().replace(day=1)
        end = start + relativedelta(months=1, days=-1)

        qs = ShopPayment.objects.filter(
            shop_id__in=shops_id,
            date__range=(start, end)
        ).exclude(
            paid_at__isnull=True
        ).values("shop_id", "payment_method").annotate(
            total_amount=Coalesce(Sum("amount"), 0)
        ).values_list("shop_id", "payment_method", "total_amount")

        current_month = defaultdict(list)
        for shop_id, pm, ta in qs.all():
            current_month[shop_id].append([pm, ta])

        context["current_month_payment_by_shop"] = current_month

        return context


class ShopQrCode(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Shop
    permission_required = 'shop.view_shop'

    def get_queryset(self):
        return super().get_queryset().filter(
            section__area__bazaar_id__in=self.request.user.allowed_bazaar.values_list('id', flat=True)
        )

    def render_to_response(self, context, **response_kwargs):
        obj = self.object  # type: Shop

        return render_qr_png_file(obj.qr_image_file)


class ShopImportView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    TITLE = _("Magazinlarni import qilish")

    model = Bazaar
    permission_required = 'shop.change_shop'
    template_name = 'shop/import.j2'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        step = 1
        try:
            step = int(request.POST.get('step', 1))
        except:
            pass

        if step == 1:
            try:
                file = request.FILES.get('file')
                workbook = load_workbook(file)
                sheet = workbook.active

                shop_wrong, duplicates, shop_skip, shop_insert, shop_update, shop_delete, import_list, can_save = self.check(
                    sheet.iter_rows(values_only=True), skip_first_line=True
                )
            except Exception as e:
                messages.error(request, str(e))
                return redirect("shop:import", self.object.id)
        elif step == 2:
            shop_list, is_delete = [], False
            try:
                shop_list = json.loads(base64.b64decode(request.POST.get('data', "")))
                is_delete = request.POST.get('delete', "0") == "1"
            except:
                step = 1
                pass

            if step == 2:
                try:
                    area = Area.objects.order_by('id').filter(bazaar_id=self.object.id).first()
                    if not area:
                        raise Area.DoesNotExist
                except Area.DoesNotExist:
                    messages.error(self.request, _("Ushbu bozorga blok kiritilmagan"))
                    return redirect("shop:import", self.object.id)

                try:
                    section = Section.objects.order_by('id').filter(area_id=area.id).first()
                    if not section:
                        raise Section.DoesNotExist
                except Section.DoesNotExist:
                    messages.error(self.request, _("Ushbu bozorga bo'lim kiritilmagan"))
                    return redirect("shop:import", self.object.id)

                with transaction.atomic():
                    shop_wrong, duplicates, shop_skip, shop_insert, shop_update, shop_delete, import_list, can_save = self.check(
                        shop_list, in_transaction=True
                    )
                    if can_save:
                        if shop_insert:
                            insert_objs = [Shop(
                                section_id=section.id,
                                number=n,
                                owner=o,
                                rent_price=p,
                                is_active=True,
                            ) for n, o, p in shop_insert]
                            Shop.objects.bulk_create(insert_objs, batch_size=10)

                        if shop_update:
                            for shop, o, p in shop_update:
                                shop.owner = o
                                shop.rent_price = p
                                shop.is_active = True
                                shop.save()

                                # Shop.objects.bulk_update(update_objs, fields=['owner', 'rent_price', 'is_active'], batch_size=10)

                        if shop_delete:
                            qs = Shop.objects.filter(
                                id__in=[row.id for row in shop_delete]
                            )
                            if is_delete:
                                qs.delete()
                            else:
                                qs.update(is_active=False)

                        messages.success(request, _("Muvaffaqiyatli import qilindi."))
                        return redirect("shop:list", self.object.id)
                    else:
                        # Agar saqlab bo'lmasa, userga ko'rsatamiz
                        step = 1

        context = self.get_context_data(object=self.object)

        context["step"] = step
        context["shop_wrong"] = shop_wrong
        context["duplicates"] = duplicates
        context["shop_insert"] = shop_insert
        context["shop_update"] = shop_update
        context["shop_delete"] = shop_delete
        context["shop_skip"] = shop_skip
        context["import_data"] = base64.b64encode(json.dumps(import_list).encode('utf-8')).decode('utf-8')
        context["can_save"] = can_save

        return self.render_to_response(context)

    def check(self, shop_list, *, in_transaction=False, skip_first_line=False):
        shop_wrong, duplicates, shop_skip, shop_insert, shop_update, shop_delete = [], [], [], [], [], []
        shop_processed, import_list = set(), []

        qs = Shop.objects.filter(
            section__area__bazaar_id=self.object.id
        )
        if in_transaction:
            qs = qs.select_for_update()

        shop_by_number = {row.number: row for row in qs.all()}

        for n, row in enumerate(shop_list):
            if skip_first_line and n == 0:  # Skip header
                continue

            if len(row) < 3:
                raise Exception(_("Faylda kamida 3 ta ustun bo'lishi lozim"))

            number, owner, price = map(lambda s: s.strip(), map(str, row[:3]))
            import_list.append((number, owner, price))

            if not re.match(Shop.NUMBER_PATTERN, number) or not re.match("^[0-9]+$", price):
                shop_wrong.append((number, owner, price))
                continue

            price = int(price)
            owner = re.sub(r'[\s\u00A0]+', ' ', owner).strip()

            if number in shop_processed:
                duplicates.append((number, owner, price))
                continue

            shop_processed.add(number)

            if number not in shop_by_number:
                # Agar DB da mavjud bo'lmasa
                shop_insert.append((number, owner, price))
            else:
                shop = shop_by_number[number]
                del shop_by_number[number]
                if not shop.owner:
                    shop.owner = ""

                if shop.owner.strip().lower() == owner.lower() and shop.rent_price == price:
                    shop_skip.append(shop)
                else:
                    shop_update.append((shop, owner, price))

        shop_delete.extend(shop_by_number.values())
        can_save = not shop_wrong and not duplicates and (shop_insert or shop_update or shop_delete)

        return shop_wrong, duplicates, shop_skip, shop_insert, shop_update, shop_delete, import_list, can_save
